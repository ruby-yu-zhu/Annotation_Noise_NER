# -*- coding: utf-8 -*-
# @Time : 2019/6/2 上午10:55
# @Author : Scofield Phil
# @FileName: BiLSTM_CRFs.py
# @Project: sequence-lableing-vex

import math, os
from engines.utils import get_test_entities_and_sentences,classify_metrics
import numpy as np
import tensorflow as tf
import pandas as pd
import time
from itertools import chain
import xlwt,xlrd
import openpyxl
import tensorflow_hub as hub



tf.logging.set_verbosity(tf.logging.ERROR)
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

class BiLSTM_CRFs(object):
    def __init__(self, configs, logger, dataManager):
        os.environ['CUDA_VISIBLE_DEVICES'] = configs.CUDA_VISIBLE_DEVICES

        self.configs = configs
        self.logger = logger
        self.logdir = configs.log_dir
        self.dataManager = dataManager
        self.measuring_metrics = configs.measuring_metrics

        self.test_file = configs.datasets_fold + "/" + configs.test_file

        if configs.mode == "train":
            self.is_training = True
        else:
            self.is_training = False

        self.checkpoint_name = configs.checkpoint_name
        self.checkpoints_dir = configs.checkpoints_dir
        self.output_test_file = configs.output_test_file

        self.biderectional = configs.biderectional
        self.cell_type = configs.cell_type
        self.num_layers = configs.encoder_layers

        self.learning_rate = configs.learning_rate
        self.dropout_rate = configs.dropout
        self.batch_size = configs.batch_size

        self.emb_dim = configs.embedding_dim
        self.hidden_dim = configs.hidden_dim

        #set char_embedding parament-------------
        self.max_char_len = configs.max_char_len
        self.char_dim = configs.char_dim
        self.char_lstm_dim = configs.char_lstm_dim
        self.chars_voab_num = dataManager.max_char_num

        self._filter_length_list = [1, 2, 3, 4, 5,6,7,8]
        self._nb_filter_list = [50, 50, 50, 50, 50,50,50,50]


        if configs.cell_type == 'LSTM':
            if self.biderectional:
                self.cell = tf.nn.rnn_cell.LSTMCell(self.hidden_dim)
            else:
                self.cell = tf.nn.rnn_cell.LSTMCell(2 * self.hidden_dim)
        else:
            if self.biderectional:
                self.cell = tf.nn.rnn_cell.GRUCell(self.hidden_dim)
            else:
                self.cell = tf.nn.rnn_cell.GRUCell(2 * self.hidden_dim)

        self.is_attention = configs.use_self_attention
        self.attention_dim = configs.attention_dim

        self.num_epochs = configs.epoch
        self.max_time_steps = configs.max_sequence_length

        self.num_tokens = dataManager.max_token_number
        self.num_classes = dataManager.max_label_number

        self.is_early_stop = configs.is_early_stop
        self.patient = configs.patient

        self.max_to_keep = configs.checkpoints_max_to_keep
        self.print_per_batch = configs.print_per_batch

        self.best_acc_dev = 0
        self.biggest_gap = 0


        if configs.optimizer == 'Adagrad':
            self.optimizer = tf.train.AdagradOptimizer(self.learning_rate)
        elif configs.optimizer == 'Adadelta':
            self.optimizer = tf.train.AdadeltaOptimizer(self.learning_rate)
        elif configs.optimizer == 'RMSprop':
            self.optimizer = tf.train.RMSPropOptimizer(self.learning_rate)
        elif configs.optimizer == 'GD':
            self.optimizer = tf.train.GradientDescentOptimizer(self.learning_rate)
        else:
            self.optimizer = tf.train.AdamOptimizer(self.learning_rate)

        self.initializer = tf.contrib.layers.xavier_initializer()
        self.global_step = tf.Variable(0, trainable=False, name="global_step", dtype=tf.int32)

        self.char_embedding = tf.get_variable("char_embedding", [self.chars_voab_num, self.char_dim],initializer=self.initializer, trainable=True)

        if configs.use_pretrained_embedding:
            embedding_matrix = dataManager.getEmbedding(configs.token_emb_dir)
            self.embedding = tf.Variable(embedding_matrix, trainable=False, name="emb", dtype=tf.float32)
        else:
            self.embedding = tf.get_variable("emb", [self.num_tokens, self.emb_dim], trainable=True,
                                             initializer=self.initializer)


        self.build()
        self.logger.info("model initialed...\n")

        self.sess = tf.Session(config=tf.ConfigProto(allow_soft_placement=True))

    def elmo_embedding(self, batch_sentence,batch_size):
        """
        to get elmo_embeddings of sentences
        :param batch_sentence: input sentences of a batch(passage_node)
        :return: a tensor after flair embedding
        """
        # batch_sentence = tf.Session().run(batch_sentence)
        module_spec = hub.load_module_spec("/home/nlp/tfhub_modules/9bb74bc86f9caffc8c47dd7b33ec4bb354d9602d/")
        elmo = hub.Module(module_spec, trainable=True)
        elmo_emb = elmo(inputs={
                                "tokens": batch_sentence,
                                "sequence_len": batch_size
                                }, signature="tokens", as_dict=True)["elmo"]
        print(elmo_emb.shape)
        return elmo_emb

    def MultiConvolutional3D(self, input_data, filter_length_list, nb_filter_list, padding='VALID', pooling='max', name='Convolutional3D'):
        """3D卷积层
                Args:
                    input_data: 4D tensor of shape=[batch_size, sent_len, word_len, char_dim]
                        in_channels is set to 1 when use Convolutional3D.
                    filter_length_list: list of int, 卷积核的长度，用于构造卷积核，在
                        Convolutional1D中，卷积核shape=[filter_length, in_width, in_channels, nb_filters]
                    nb_filter_list: list of int, 卷积核数量
                    padding: 默认'VALID'，暂时不支持设成'SAME'
                """
        assert padding in ('VALID'), 'Unknow padding %s' % padding
        # assert padding in ('VALID', 'SAME'), 'Unknow padding %s' % padding

        # expand dim
        char_dim = int(input_data.get_shape()[-1])  # char的维度
        input_data = tf.expand_dims(input_data, -1)  # shape=[x, x, x, 1]
        pooling_outpouts = []
        for i in range(len(filter_length_list)):
            filter_length = filter_length_list[i]
            nb_filter = nb_filter_list[i]
            with tf.variable_scope('%s_%d' % (name, filter_length)) as scope:
                # shape= [batch_size, sent_len-filter_length+1, word_len, 1, nb_filters]
                conv_output = tf.contrib.layers.conv3d(
                    inputs=input_data,
                    num_outputs=nb_filter,
                    kernel_size=[1, filter_length, char_dim],
                    padding=padding)
                # output's shape=[batch_size, new_height, 1, nb_filters]
                act_output = tf.nn.relu(conv_output)
                # max pooling，shape = [batch_size, sent_len, nb_filters]
                if pooling == 'max':
                    pooling_output = tf.reduce_max(tf.squeeze(act_output, [-2]), 2)
                elif pooling == 'mean':
                    pooling_output = tf.reduce_mean(tf.squeeze(act_output, [-2]), 2)
                else:
                    raise Exception('pooling must in (max, mean)!')
                pooling_outpouts.append(pooling_output)

                scope.reuse_variables()
        # [batch_size, sent_len, sum(nb_filter_list]
        output = tf.concat(pooling_outpouts, axis=-1)
        return output

    def BiLSTM(self,inputs_emb):
        lstm_cell_fw = self.cell
        lstm_cell_bw = self.cell
        if self.is_training:
            lstm_cell_fw = tf.nn.rnn_cell.DropoutWrapper(lstm_cell_fw, output_keep_prob=(1 - self.dropout_rate))
            lstm_cell_bw = tf.nn.rnn_cell.DropoutWrapper(lstm_cell_bw, output_keep_prob=(1 - self.dropout_rate))

        lstm_cell_fw = tf.nn.rnn_cell.MultiRNNCell([lstm_cell_fw] * self.num_layers)
        lstm_cell_bw = tf.nn.rnn_cell.MultiRNNCell([lstm_cell_bw] * self.num_layers)
        outputs, _, _ = tf.contrib.rnn.static_bidirectional_rnn(
            lstm_cell_fw,
            lstm_cell_bw,
            inputs_emb,
            dtype=tf.float32,
        )
        # outputs: list_steps[batch, 2*dim]
        outputs = tf.concat(outputs, 1)
        outputs = tf.reshape(outputs, [self.batch_size, self.max_time_steps, self.hidden_dim * 2])
        return outputs

    def self_attention(self,input_emb,input_mask):
        H1 = tf.reshape(input_emb, [-1, self.hidden_dim * 2])
        W_a1 = tf.get_variable("W_a1", shape=[self.hidden_dim * 2, self.attention_dim],
                               initializer=self.initializer, trainable=True)
        u1 = tf.matmul(H1, W_a1)

        H2 = tf.reshape(tf.identity(input_emb), [-1, self.hidden_dim * 2])
        W_a2 = tf.get_variable("W_a2", shape=[self.hidden_dim * 2, self.attention_dim],
                               initializer=self.initializer, trainable=True)
        u2 = tf.matmul(H2, W_a2)

        u1 = tf.reshape(u1, [self.batch_size, self.max_time_steps, self.hidden_dim * 2])
        u2 = tf.reshape(u2, [self.batch_size, self.max_time_steps, self.hidden_dim * 2])
        u = tf.matmul(u1, u2, transpose_b=True)

        # Array of weights for each time step
        A = tf.nn.softmax(u, name="attention")
        outputs = tf.matmul(A, tf.reshape(tf.identity(input_emb),
                                          [self.batch_size, self.max_time_steps, self.hidden_dim * 2]))
        output_mask =tf.multiply(outputs,tf.expand_dims(input_mask, axis=-1))
        return output_mask

    def separate_context_feature(self,output):
        forward_words_hiddens = output[:,0:10:]
        backward_words_hiddens = output[:,-10::]
        context_feature = forward_words_hiddens+backward_words_hiddens
        sum_context_feature = tf.reduce_sum(context_feature,1)#[batch_size,attention_dim]
        return sum_context_feature

    def separate_mention_feature(self,output):
        mention_feature = output[:,10:15:]
        mention_char_feature = self.inputs_char_emb[:,10:15:]
        mention_feature = tf.add(mention_feature,mention_char_feature)
        sum_mention_feature = tf.reduce_sum(mention_feature,1)#[batch_size,attention_dim]
        return sum_mention_feature

    def multi_category_focal_loss(self,y_pred,y_true):
        epsilon = 1.e-7
        gamma = 2.0
        # alpha = tf.constant([[2],[1],[1],[1],[1]], dtype=tf.float32)
        alpha = tf.constant([[1], [1], [1], [2]], dtype=tf.float32)#constract with y_labels

        y_true = tf.cast(y_true, tf.float32)
        y_pred = tf.clip_by_value(y_pred, epsilon, 1. - epsilon)
        y_t = tf.multiply(y_true, y_pred) + tf.multiply(1 - y_true, 1 - y_pred)
        ce = -tf.log(y_t)
        weight = tf.pow(tf.subtract(1., y_t), gamma)
        fl = tf.matmul(tf.multiply(weight, ce), alpha)
        return fl

    def combain_pre(self,context_feature,mention_feature):
        with tf.variable_scope("label"):
            combain_feature = tf.concat([context_feature,mention_feature],-1)#[batch_size,attention_dim*2]
            softmax_w = tf.get_variable("softmax_w", [self.attention_dim*2, self.num_classes],
                                             initializer=self.initializer)
            softmax_b = tf.get_variable("softmax_b", [self.num_classes], initializer=self.initializer)
            logits = tf.matmul(combain_feature,softmax_w) + softmax_b
            self.y_scores = tf.nn.softmax(logits, axis=-1)
            with tf.name_scope("loss"):
                losses = self.multi_category_focal_loss(self.y_scores, self.targets)
                self.combain_loss = tf.reduce_mean(losses, name="task_loss")

            with tf.name_scope("accuracy"):
                self.y_pred = tf.argmax(self.y_scores, 1, name="predictions")

    def mention_distinctor(self,mention_features):
        with tf.variable_scope("mention"):
            softmax_w = tf.get_variable("softmax_w", [self.attention_dim, self.num_classes],
                                             initializer=self.initializer)
            softmax_b = tf.get_variable("softmax_b", [self.num_classes], initializer=self.initializer)
            logits = tf.matmul(mention_features,softmax_w) + softmax_b
            self.m_scores = tf.nn.softmax(logits, axis=-1)
            with tf.name_scope("loss"):
                losses = self.multi_category_focal_loss(self.m_scores, self.targets)
                self.m_loss = tf.reduce_mean(losses, name="task_loss")

            with tf.name_scope("accuracy"):
                self.m_pred = tf.argmax(self.m_scores, 1, name="predictions")

    def context_distinctor(self,context_feature):
        with tf.variable_scope("mention",reuse = True):
            softmax_w = tf.get_variable("softmax_w", [self.attention_dim, self.num_classes],
                                             initializer=self.initializer)
            softmax_b = tf.get_variable("softmax_b", [self.num_classes], initializer=self.initializer)
            logits = tf.matmul(context_feature,softmax_w) + softmax_b
            self.c_scores = tf.nn.softmax(logits, axis=-1)
            with tf.name_scope("context_loss"):
                losses = self.multi_category_focal_loss(self.c_scores, self.targets)
                self.c_loss = tf.reduce_mean(losses, name="task_loss")

            with tf.name_scope("context_accuracy"):
                self.c_pred = tf.argmax(self.c_scores, 1, name="predictions")


    def build(self):
        self.inputs = tf.placeholder(tf.int64, [None, self.max_time_steps])
        self.inputs_str = tf.placeholder(tf.string, [None, None])
        self.input_mask = tf.placeholder(tf.float32, [None, self.max_time_steps])#mask padding words
        self.inputs_char = tf.placeholder(tf.int64, [None, self.max_time_steps, self.max_char_len])#[batch_size,max_seq_len,max_char_len]

        self.targets = tf.placeholder(tf.int64, [None,self.num_classes])#batch_size

        char_emb = tf.nn.embedding_lookup(self.char_embedding,self.inputs_char)#[batch_size,max_seq_len,max_char_len,char_dim]
        self.inputs_char_emb = self.MultiConvolutional3D(char_emb, self._filter_length_list,self._nb_filter_list)
        print("sentences after char_embeddings: ",self.inputs_char_emb.shape)

        self.inputs_elmo_emb = self.elmo_embedding(self.inputs_str,self.batch_size*[self.max_time_steps])#replace_elmo word_embeedings
        self.inputs_emb = tf.concat([self.inputs_elmo_emb,self.inputs_char_emb],axis=-1)
        print("sentences after concating elmo: ",self.inputs_emb.shape)

        # self.inputs_emb = tf.nn.embedding_lookup(self.embedding, self.inputs)
        self.inputs_emb = tf.transpose(self.inputs_emb, [1, 0, 2])
        self.inputs_emb = tf.reshape(self.inputs_emb, [-1, self.char_lstm_dim+self.emb_dim])
        self.inputs_emb = tf.split(self.inputs_emb, self.max_time_steps, 0)

        bilstm_ouputs = self.BiLSTM(self.inputs_emb)
        print("sentences after bilstm: ",bilstm_ouputs.shape)
        att_outputs = self.self_attention(bilstm_ouputs,self.input_mask)
        print("sentences after attentions: ",att_outputs.shape)
        context_feature = self.separate_context_feature(att_outputs)
        print("shape of context features: ",context_feature.shape)
        mention_feature = self.separate_mention_feature(att_outputs)
        print("shape of mention features: ",mention_feature.shape)
        self.combain_pre(context_feature,mention_feature)
        self.mention_distinctor(mention_feature)
        self.context_distinctor(context_feature)

        all_var_list = tf.trainable_variables()
        print(all_var_list)

        self.optimizer_n = tf.train.AdamOptimizer(
            learning_rate=self.learning_rate
        ).minimize(
            self.combain_loss,
            global_step=self.global_step
        )  # optimized all

        var_d = [var for var in all_var_list if 'mention' in var.name]
        print(var_d)
        self.optimizer_d = tf.train.AdamOptimizer(
            learning_rate=self.learning_rate
        ).minimize(
            self.m_loss + self.c_loss,
            var_list=var_d
        )  # optimized distinctor

        var_g = [var for var in all_var_list if var not in var_d]
        print(var_g)
        self.optimizer_g = tf.train.AdamOptimizer(
            learning_rate=self.learning_rate
        ).minimize(
            self.combain_loss + self.m_loss - self.c_loss,
            var_list=var_g,
            global_step=self.global_step
        )  # optimized generator


    def train(self):
        X_train_tokens, X_train_token_ids, X_train_mask, X_train_chars, y_train_ids, \
        X_dev_tokens, X_dev_token_ids, X_dev_mask, X_dev_chars, y_dev_ids = self.dataManager.getTrainingSet()
        tf.initialize_all_variables().run(session=self.sess)

        saver = tf.train.Saver(max_to_keep=self.max_to_keep)
        tf.summary.merge_all()

        num_iterations = int(math.ceil(1.0 * len(X_train_tokens) / self.batch_size))
        num_val_iterations = int(math.ceil(1.0 * len(X_dev_tokens) / self.batch_size))

        cnt = 0
        cnt_dev = 0
        unprogressed = 0
        very_start_time = time.time()
        best_at_epoch = 0
        self.logger.info("\ntraining starting" + ("+" * 20))
        for epoch in range(self.num_epochs):
            start_time = time.time()
            # shuffle train at each epoch
            sh_index = np.arange(len(X_train_tokens))
            np.random.shuffle(sh_index)
            X_train_tokens = X_train_tokens[sh_index]
            X_train_token_ids = X_train_token_ids[sh_index]
            X_train_mask = X_train_mask[sh_index]
            X_train_chars = X_train_chars[sh_index]
            y_train_ids = y_train_ids[sh_index]
            #=========pre_train=====================
            self.logger.info("\ncurrent epoch: %d" % (epoch))
            for iteration in range(num_iterations):
                X_train_tokens_batch, X_train_token_ids_batch,X_train_mask_batch,X_train_chars_batch,y_train_ids_batch = self.dataManager.nextBatch(X_train_tokens,X_train_token_ids,X_train_mask,
                                                                                                                                                    X_train_chars,y_train_ids,start_index=iteration * self.batch_size)
                # print(X_train_tokens_batch.shape)
                # print(X_train_token_ids_batch.shape)
                # print(X_train_mask_batch.shape)
                # print(X_train_chars_batch.shape)
                # print(y_train_ids_batch.shape)
                _, train_combain_loss, train_combain_pred = \
                    self.sess.run([
                        self.optimizer_n, self.combain_loss, self.y_pred
                    ],
                        feed_dict={
                            self.inputs: X_train_token_ids_batch,
                            self.inputs_str: X_train_tokens_batch,
                            self.input_mask: X_train_mask_batch,
                            self.inputs_char: X_train_chars_batch,
                            self.targets: y_train_ids_batch
                        })

                train_combain_measu = classify_metrics(y_train_ids_batch,train_combain_pred,self.measuring_metrics)

                if iteration % self.print_per_batch == 0:
                    cnt += 1
                    self.logger.info("training batch: %5d, combain loss: %.5f, combain macro-f1: %.5f"
                                     % (iteration, train_combain_loss,train_combain_measu['f1']))
            #----------for dev---------------
            dev_combain_losses = []
            dev_results = dict()
            for measu in self.measuring_metrics:
                dev_results[measu] = 0
            for iterr in range(num_val_iterations):
                cnt_dev += 1
                X_dev_tokens_batch,X_dev_token_ids_batch,X_dev_mask_batch,X_dev_chars_batch,y_dev_ids_batch = self.dataManager.nextBatch(X_dev_tokens, X_dev_token_ids, X_dev_mask,
                                                                                                                                         X_dev_chars, y_dev_ids,start_index=iterr * self.batch_size)
                dev_combain_loss, dev_combain_pred= \
                    self.sess.run([
                        self.combain_loss, self.y_pred
                    ],
                        feed_dict={
                            self.inputs: X_dev_token_ids_batch,
                            self.inputs_str: X_dev_tokens_batch,
                            self.input_mask: X_dev_mask_batch,
                            self.inputs_char: X_dev_chars_batch,
                            self.targets: y_dev_ids_batch
                        })
                dev_combain_losses.append(dev_combain_loss)
                dev_combain_measu = classify_metrics(y_dev_ids_batch,dev_combain_pred,self.measuring_metrics)
                for k, v in dev_combain_measu.items():
                    dev_results[k] += v

            averg_combain_loss = np.array(dev_combain_losses).mean()
            time_span = (time.time() - start_time) / 60
            for k, v in dev_results.items():
                dev_results[k] /= num_val_iterations

            self.logger.info("time consumption:%.2f(min),  validation combain loss: %.5f, validation combain macro-f1: %.5f" %
                             (time_span, averg_combain_loss, dev_results['f1']))


        #===========for get_more_mention_feature========================
        self.logger.info("\ntraining starting" + ("+" * 20))
        for epoch in range(self.num_epochs):
            start_time = time.time()
            # shuffle train at each epoch
            sh_index = np.arange(len(X_train_tokens))
            np.random.shuffle(sh_index)
            X_train_tokens = X_train_tokens[sh_index]
            X_train_token_ids = X_train_token_ids[sh_index]
            X_train_mask = X_train_mask[sh_index]
            X_train_chars = X_train_chars[sh_index]
            y_train_ids = y_train_ids[sh_index]

            self.logger.info("\ncurrent epoch: %d" % (epoch))
            for iteration in range(num_iterations):
                X_train_tokens_batch, X_train_token_ids_batch, X_train_mask_batch, X_train_chars_batch, y_train_ids_batch = self.dataManager.nextBatch(
                    X_train_tokens, X_train_token_ids, X_train_mask,
                    X_train_chars, y_train_ids, start_index=iteration * self.batch_size)
                # print(X_train_tokens_batch.shape)
                # print(X_train_token_ids_batch.shape)
                # print(X_train_mask_batch.shape)
                # print(X_train_chars_batch.shape)
                # print(y_train_ids_batch.shape)
                _, _, _= \
                    self.sess.run([
                        self.optimizer_d, self.m_loss, self.m_pred
                    ],
                        feed_dict={
                            self.inputs: X_train_token_ids_batch,
                            self.inputs_str: X_train_tokens_batch,
                            self.input_mask: X_train_mask_batch,
                            self.inputs_char: X_train_chars_batch,
                            self.targets: y_train_ids_batch
                        })

                _, train_context_loss, train_context_pred,train_mention_loss,train_mention_pred = \
                    self.sess.run([
                        self.optimizer_g, self.c_loss, self.c_pred,self.m_loss,self.m_pred
                    ],
                        feed_dict={
                            self.inputs: X_train_token_ids_batch,
                            self.inputs_str: X_train_tokens_batch,
                            self.input_mask: X_train_mask_batch,
                            self.inputs_char: X_train_chars_batch,
                            self.targets: y_train_ids_batch
                        })

                train_context_measu = classify_metrics(y_train_ids_batch,train_context_pred,self.measuring_metrics)
                train_mention_measu = classify_metrics(y_train_ids_batch,train_mention_pred,self.measuring_metrics)

                if iteration % self.print_per_batch == 0:
                    cnt += 1
                    self.logger.info(
                        "training batch: %5d, context loss: %.5f, context macro-f1: %.5f, mention loss: %.5f, mention macro-f1: %.5f" % (iteration, train_context_loss, train_context_measu['f1'],train_mention_loss,train_mention_measu['f1']))
            # ----------for dev---------------
            dev_context_losses = []
            dev_mention_losses = []
            dev_combain_losses = []
            dev_combain_results = dict()
            dev_context_results = dict()
            dev_mention_results = dict()
            for measu in self.measuring_metrics:
                dev_combain_results[measu] = 0
                dev_context_results[measu] = 0
                dev_mention_results[measu] = 0
            for iterr in range(num_val_iterations):
                cnt_dev += 1
                X_dev_tokens_batch, X_dev_token_ids_batch, X_dev_mask_batch, X_dev_chars_batch, y_dev_ids_batch = self.dataManager.nextBatch(
                    X_dev_tokens, X_dev_token_ids, X_dev_mask,
                    X_dev_chars, y_dev_ids, start_index=iterr * self.batch_size)
                dev_combain_loss,dev_combain_pred,dev_context_loss, dev_context_pred,dev_mention_loss,dev_mention_pred= \
                    self.sess.run([
                        self.combain_loss,self.y_pred,self.c_loss,self.c_pred,self.m_loss,self.m_pred
                    ],
                        feed_dict={
                            self.inputs: X_dev_token_ids_batch,
                            self.inputs_str: X_dev_tokens_batch,
                            self.input_mask: X_dev_mask_batch,
                            self.inputs_char: X_dev_chars_batch,
                            self.targets: y_dev_ids_batch
                        })


                dev_context_losses.append(dev_context_loss)
                dev_mention_losses.append(dev_mention_loss)
                dev_combain_losses.append(dev_combain_loss)
                dev_combain_measu = classify_metrics(y_dev_ids_batch,dev_combain_pred,self.measuring_metrics)
                dev_context_measu = classify_metrics(y_dev_ids_batch,dev_context_pred,self.measuring_metrics)
                dev_mention_measu = classify_metrics(y_dev_ids_batch,dev_mention_pred,self.measuring_metrics)

                for k, v in dev_combain_measu.items():
                    dev_combain_results[k] += v
                for k, v in dev_context_measu.items():
                    dev_context_results[k] += v
                for k, v in dev_mention_measu.items():
                    dev_mention_results[k] += v

            time_span = (time.time() - start_time) / 60
            for k, v in dev_combain_results.items():
                dev_combain_results[k] /= num_val_iterations
            for k, v in dev_context_results.items():
                dev_context_results[k] /= num_val_iterations
            for k, v in dev_mention_results.items():
                dev_mention_results[k] /= num_val_iterations

            dev_acc_avg = dev_mention_results['f1']
            acc_gap = dev_mention_results['f1'] - dev_context_results['f1']
            self.logger.info("time consumption:%.2f(min), validation combain loss: %.5f, validation combain macro-f1: %.5f, validation context loss: %.5f, validation context macro-f1: %.5f, validation mention loss: %.5f, validation mention macro-f1: %.5f, differ of context and mention is: %.5f" %
                             (time_span, np.array(dev_combain_losses).mean(), dev_combain_results['f1'], np.array(dev_context_losses).mean(), dev_context_results['f1'], np.array(dev_mention_losses).mean(), dev_mention_results['f1'],acc_gap))


            if dev_acc_avg > self.best_acc_dev or acc_gap > self.biggest_gap:
                unprogressed = 0
                self.best_acc_dev = dev_acc_avg
                self.biggest_gap = acc_gap
                best_at_epoch = epoch
                saver.save(self.sess, self.checkpoints_dir + "/" + self.checkpoint_name, global_step=self.global_step)
                self.logger.info("saved the new best model with combain macro-f1 is: %.3f, with context macro-f1 is: %.3f, with mention macro-f1 is: %.3f, differ of context and mention is: %.3f" % (self.best_acc_dev,dev_context_results['f1'],dev_mention_results['f1'],self.biggest_gap))
            else:
                unprogressed += 1

            if self.is_early_stop:
                if unprogressed >= self.patient:
                    self.logger.info("early stopped, no progress obtained within %d epochs" % self.patient)
                    self.logger.info("overall best recall is %f at %d epoch" % (self.best_acc_dev, best_at_epoch))
                    self.logger.info(
                        "total training time consumption: %.3f(min)" % ((time.time() - very_start_time) / 60))
                    self.sess.close()
                    return
        self.logger.info("overall best f1 is %f at %d epoch" % (self.best_acc_dev, best_at_epoch))
        self.logger.info("total training time consumption: %.3f(min)" % ((time.time() - very_start_time) / 60))

        self.sess.close()

    def test(self):
        X_test_tokens,X_test_ids,X_test_mask,X_test_chars,y_test_ids = self.dataManager.getTestingSet()
        # _, _, Y_test_str = self.dataManager.getTestingrealY_str()

        num_iterations = int(math.ceil(1.0 * len(X_test_tokens) / self.batch_size))
        self.logger.info("total number of testing iterations: " + str(num_iterations))

        self.logger.info("loading model parameter\n")
        tf.initialize_all_variables().run(session=self.sess)
        saver = tf.train.Saver()
        saver.restore(self.sess, tf.train.latest_checkpoint(self.checkpoints_dir))

        self.logger.info("\ntesting starting" + ("+" * 20))
        start_time = time.time()
        test_y_pres = []
        test_c_pres = []
        test_combain_results = dict()
        test_context_results = dict()
        for measu in self.measuring_metrics:
            test_combain_results[measu] = 0
            test_context_results[measu] = 0
        test_c_M_score = []
        test_c_D_score = []
        test_c_C_score = []
        test_c_O_score = []
        X_test_sentences=[]
        X_test_entities = []
        test_y_real = []
        for i in range(num_iterations):
            self.logger.info("batch: " + str(i + 1))
            X_test_tokens_batch,X_test_ids_batch,X_test_mask_batch,X_test_chars_batch,y_test_ids_batch \
                = self.dataManager.nextBatch(X_test_tokens,X_test_ids,X_test_mask,X_test_chars,y_test_ids,start_index=i*self.batch_size)
            # print(X_test_tokens_batch.shape,X_test_tokens_batch[0])
            # print(X_test_ids_batch.shape,X_test_ids_batch[0])
            # print(X_test_mask_batch.shape,X_test_mask_batch[0])
            # print(X_test_chars_batch.shape,X_test_chars_batch[0])
            # print((y_test_ids_batch.shape,y_test_ids_batch[0]))

            test_y_pred,test_c_pred,test_c_score = \
                self.sess.run([
                    self.y_pred,self.m_pred,self.m_scores
                ],
                    feed_dict={
                        self.inputs: X_test_ids_batch,
                        self.inputs_str:X_test_tokens_batch,
                        self.input_mask: X_test_mask_batch,
                        self.inputs_char: X_test_chars_batch,
                        self.targets: y_test_ids_batch
                    })
            y_pred_labels = [self.dataManager.id2label[y_id] for y_id in test_y_pred]
            y_real_labels = [self.dataManager.id2label[np.nonzero(y_id_list)[0][0]] for y_id_list in y_test_ids_batch]
            test_c_pre = [self.dataManager.id2label[y_id] for y_id in test_c_pred]
            test_y_pres+=y_pred_labels
            test_y_real+=y_real_labels
            test_c_pres+=test_c_pre
            test_combain_measu = classify_metrics(y_test_ids_batch,test_y_pred,self.measuring_metrics)
            test_context_measu = classify_metrics(y_test_ids_batch,test_c_pred,self.measuring_metrics)
            for k, v in test_combain_measu.items():
                test_combain_results[k] += v
            for k, v in test_context_measu.items():
                test_context_results[k] += v
            for score_list in test_c_score:
                test_c_O_score.append(score_list[0])#LOC
                test_c_M_score.append(score_list[1])#PER
                test_c_D_score.append(score_list[2])#ORG
                test_c_C_score.append(score_list[3])#MISC
            X_test_sentence,X_test_entity = get_test_entities_and_sentences(X_test_tokens_batch)
            X_test_sentences+=X_test_sentence
            X_test_entities+=X_test_entity

        X_test_sentences = X_test_sentences[0:len(X_test_tokens)]
        X_test_entities = X_test_entities[0:len(X_test_tokens)]
        test_c_pres = test_c_pres[0:len(X_test_tokens)]
        test_y_real = test_y_real[0:len(X_test_tokens)]
        test_c_M_score = test_c_M_score[0:len(X_test_tokens)]
        test_c_D_score = test_c_D_score[0:len(X_test_tokens)]
        test_c_C_score = test_c_C_score[0:len(X_test_tokens)]
        test_c_O_score = test_c_O_score[0:len(X_test_tokens)]

        sens_ids = xlrd.open_workbook(self.test_file,'rb').sheet_by_name('Sheet').col_values(2)
        entities_ids = xlrd.open_workbook(self.test_file,'rb').sheet_by_name('Sheet').col_values(3)

        time_span = (time.time() - start_time) / 60
        for k, v in test_combain_results.items():
            test_combain_results[k] /= num_iterations
        test_reg_str = ''
        for k, v in test_context_results.items():
            test_context_results[k] /= num_iterations
            test_reg_str += (k + ": %.3f " % test_context_results[k])
        self.logger.info("time consumption:%.2f(min), test combain macro-f1: %.5f; test mention %s" %
                         (time_span, test_combain_results['f1'], test_reg_str))

        def pick_sentences(sentences,entities,types_pred,types_real):
            wrong_sentences = []
            wrong_entities = []
            wrong_types_pred = []
            wrong_types_real = []
            true_sentences = []
            true_entities = []
            true_types_pred = []
            true_types_real = []
            for i in range(len(sentences)):
                if types_pred[i] != types_real[i]:
                    wrong_sentences.append(sentences[i])
                    wrong_entities.append(entities[i])
                    wrong_types_pred.append(types_pred[i])
                    wrong_types_real.append(types_real[i])
                else:
                    true_sentences.append(sentences[i])
                    true_entities.append(entities[i])
                    true_types_pred.append(types_pred[i])
                    true_types_real.append(types_real[i])
            return wrong_sentences,wrong_entities,wrong_types_pred,wrong_types_real,\
                   true_sentences,true_entities,true_types_pred,true_types_real


        def output_test_results(sens_ids,entities_ids,sentences,entities,types_pred,types_real,M_score,D_score,C_score,O_score,output_filename,is_context=False):
            xls = openpyxl.Workbook()
            sheet = xls.get_sheet_by_name('Sheet')
            for i in range(len(sentences)):
                sheet.cell(row=i + 1, column=1, value=sens_ids[i])
                sheet.cell(row=i+1,column=2,value=' '+sentences[i])
                sheet.cell(row=i+1,column=3,value=entities[i])
                sheet.cell(row=i+1,column=4,value=types_pred[i])
                sheet.cell(row=i+1,column=5,value=types_real[i])
                sheet.cell(row=i + 1, column=10, value=entities_ids[i])
                if is_context:
                    sheet.cell(row=i + 1, column=6, value=M_score[i])
                    sheet.cell(row=i + 1, column=7, value=D_score[i])
                    sheet.cell(row=i + 1, column=8, value=C_score[i])
                    sheet.cell(row=i + 1, column=9, value=O_score[i])

            xls.save(output_filename)

        # wrong_sentences, wrong_entities, wrong_types_pred, wrong_types_real, \
        # true_sentences, true_entities, true_types_pred, true_types_real = pick_sentences(X_test_sentences,X_test_entities,test_y_pres,test_y_real)
        # output_test_results(wrong_sentences, wrong_entities, wrong_types_pred, wrong_types_real,test_c_P_score,test_c_M_score,test_c_D_score,test_c_C_score,test_c_O_score,'wrong_comabin_test_results_output.xlsx')
        # output_test_results(true_sentences, true_entities, true_types_pred, true_types_real,test_c_P_score,test_c_M_score,test_c_D_score,test_c_C_score,test_c_O_score,
        #                     'true_combain_test_results_output.xlsx')

        output_test_results(sens_ids,entities_ids,X_test_sentences,X_test_entities, test_c_pres, test_y_real,
                            test_c_M_score, test_c_D_score, test_c_C_score, test_c_O_score,
                            self.output_test_file,is_context=True)

        self.sess.close()